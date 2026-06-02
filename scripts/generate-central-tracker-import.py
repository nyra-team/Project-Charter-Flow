"""
Generator: Central Project Tracker.xlsx → idempotent SQL.

Reads ../../../Central Project Tracker.xlsx (Project-wise Summary +
Project Tracker sheets), builds a SQL file that creates one pmo_projects
row + one pmo_charters shell + N pmo_milestones rows per project. Skips
projects already present in pmo_projects (by exact name).

Output: apps/pmo/scripts/central-tracker-import.sql

Run:
    cd apps/pmo
    ../../.venv-xlsx/bin/python3 scripts/generate-central-tracker-import.py
    psql "$DATABASE_URL" -f scripts/central-tracker-import.sql

Each project is wrapped in its own DO block so a single bad row doesn't
abort the whole batch — pg's plpgsql `IF NOT EXISTS` guard makes re-runs
no-ops, so it's safe to run repeatedly.

Naming-mismatch tolerance: the Project Tracker sheet uses slight name
variants ("Catalog Based Procurement" vs "Catalog based Procurement",
"OT Security - Project 1 OT EDR" vs "OT Security"). We normalize on
lowercase + collapse whitespace + strip punctuation for the milestone
match, then attach via the canonical summary name.
"""

from __future__ import annotations

import os
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]  # repo root
XLSX_PATH = ROOT / "Central Project Tracker.xlsx"
OUT_PATH = ROOT / "apps/pmo/scripts/central-tracker-import.sql"

# ─── Helpers ────────────────────────────────────────────────────────────────

def sql_str(v):
    """Escape a Python value as a SQL string literal (or NULL)."""
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, datetime):
        return f"'{v.strftime('%Y-%m-%d')}'"
    s = str(v).replace("\\", "\\\\").replace("'", "''")
    return f"'{s}'"

def fmt_date(v):
    """SQL string literal for a YYYY-MM-DD date column, or NULL."""
    if v is None:
        return "NULL"
    if isinstance(v, datetime):
        return f"'{v.strftime('%Y-%m-%d')}'"
    return sql_str(str(v))

def norm(name: str) -> str:
    """Collapse for fuzzy project-name matching across sheets."""
    if not name:
        return ""
    s = name.lower()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

STATUS_MAP = {
    "completed": "completed",
    "in progress": "in_progress",
    "under progress": "in_progress",
    "pending": "not_started",
    "to be started": "not_started",
    "not started": "not_started",
    "na": "not_started",
    "-": "not_started",
    None: "not_started",
    "": "not_started",
}

RAG_BY_STATUS = {
    "completed": "green",
    "in_progress": "amber",
    "not_started": "green",
}

def map_status(s):
    if s is None:
        return "not_started"
    return STATUS_MAP.get(str(s).strip().lower(), "not_started")

def project_rag(project_status: str | None, milestones: list[dict]) -> str:
    """RAG for the project row — overall On Track→green, else derive."""
    if project_status and project_status.strip().lower() == "on track":
        return "green"
    return "amber"

def project_progress(milestones: list[dict]) -> int:
    """Percent of milestones in 'completed' state."""
    if not milestones:
        return 0
    done = sum(1 for m in milestones if map_status(m["status"]) == "completed")
    return round(100 * done / len(milestones))

def safe_dates(start, end, today: str = "2026-05-28") -> tuple[str | None, str | None]:
    """Source has some inverted (end<start) date pairs — preserve them as-is
    but make sure neither column ends up empty when only one was set."""
    s = start.strftime("%Y-%m-%d") if isinstance(start, datetime) else None
    e = end.strftime("%Y-%m-%d") if isinstance(end, datetime) else None
    return s, e

# ─── Read source ────────────────────────────────────────────────────────────

if not XLSX_PATH.exists():
    print(f"ERROR: not found: {XLSX_PATH}", file=sys.stderr)
    sys.exit(1)

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws_sum = wb["Project-wise Summary"]
ws_tr = wb["Project Tracker"]

# Summary: header row 12 (0-indexed), data from 13
summary_rows = list(ws_sum.iter_rows(values_only=True))[13:]
projects = []
for r in summary_rows:
    name = r[3]
    if not name:
        continue
    projects.append({
        "department": r[0],
        "function": r[1],
        "domain": r[2],
        "name": str(name).strip(),
        "description": r[4],
        "owner": r[5],
        "approved_budget": r[6],
        "actual_cost": r[7],
        "cost_variance": r[8],
        "planned_start": r[9],
        "planned_end": r[10],
        "actual_start": r[11],
        "actual_end": r[12],
        "estimated_end": r[13],
        "schedule_variance": r[14],
        "progress_raw": r[15],
        "comments": r[16],
        "risk": r[17],
        "risk_desc": r[18],
        "mitigation": r[19],
    })

# Tracker: header row 8 (0-indexed), data from 9
tracker_rows = list(ws_tr.iter_rows(values_only=True))[9:]
milestones_by_name: dict[str, list[dict]] = {}
project_status_by_name: dict[str, str] = {}
for r in tracker_rows:
    pname = r[3]
    mname = r[8]
    if not pname or not mname:
        continue
    pname = str(pname).strip()
    milestones_by_name.setdefault(pname, []).append({
        "phase": r[7],
        "name": str(mname).strip(),
        "start": r[9],
        "target": r[10],
        "owner": r[11],
        "status": r[12],
        "comments": r[13],
        "risks": r[14],
    })
    if r[6]:
        project_status_by_name.setdefault(pname, str(r[6]).strip())

# Build a normalized index so the summary's "Catalog based Procurement"
# matches the tracker's "Catalog Based Procurement", etc.
tracker_norm_index = {norm(k): k for k in milestones_by_name.keys()}

def milestones_for(summary_name: str) -> list[dict]:
    """Look up milestones from tracker — exact, then normalized."""
    if summary_name in milestones_by_name:
        return milestones_by_name[summary_name]
    n = norm(summary_name)
    if n in tracker_norm_index:
        return milestones_by_name[tracker_norm_index[n]]
    # Loose contains-match (e.g. "OT Security" vs "OT Security - Project 1 …")
    matches = [orig for k, orig in tracker_norm_index.items() if n in k or k in n]
    if matches:
        # Combine milestones from all loose matches
        combined = []
        for m in matches:
            combined.extend(milestones_by_name[m])
        return combined
    return []

# ─── Generate SQL ──────────────────────────────────────────────────────────

out_lines: list[str] = []
out_lines.append(
    "-- ============================================================================\n"
    "-- Central Project Tracker import — auto-generated by\n"
    "-- apps/pmo/scripts/generate-central-tracker-import.py\n"
    "-- DO NOT EDIT BY HAND. Re-run the generator if the source xlsx changes.\n"
    "--\n"
    "-- Apply with:\n"
    "--   cd apps/pmo\n"
    "--   psql \"$DATABASE_URL\" -f scripts/central-tracker-import.sql\n"
    "--\n"
    "-- Each project is wrapped in its own DO block with IF NOT EXISTS guard,\n"
    "-- so re-running this file is a no-op for projects already present.\n"
    "-- One BEGIN/COMMIT wraps the whole batch so a single bad row doesn't\n"
    "-- leave the DB half-populated.\n"
    "-- ============================================================================\n\n"
    "BEGIN;\n"
)

inserted_count = 0
total_milestones = 0
skipped_projects: list[str] = []

for p in projects:
    name = p["name"]
    ms = milestones_for(name)

    # Project description: build from available metadata
    description_parts = [
        f"Imported from \"Central Project Tracker.xlsx\".",
        "",
        f"Department: {p['department'] or '—'}",
        f"Function: {p['function'] or '—'}",
        f"Domain: {p['domain'] or '—'}",
        f"Owner: {p['owner'] or '—'}",
    ]
    if p["description"]:
        description_parts.extend(["", str(p["description"])])
    if p["comments"]:
        description_parts.extend(["", f"Comments: {p['comments']}"])
    if p["risk"] or p["risk_desc"] or p["mitigation"]:
        description_parts.extend([
            "",
            f"Risk/Dependency/Issue: {p['risk'] or '—'}",
            f"Description: {p['risk_desc'] or '—'}",
            f"Mitigation: {p['mitigation'] or '—'}",
        ])
    description = "\n".join(description_parts)

    start_date, end_date = safe_dates(p["planned_start"], p["planned_end"])
    proj_status = project_status_by_name.get(name) or "On Track"
    proj_rag = project_rag(proj_status, ms)
    progress = project_progress(ms)

    charter_scope = (
        f"## Scope\n\nImported from the Central Project Tracker. "
        f"{ms.__len__()} milestone(s) tracked across the project lifecycle.\n\n"
        f"## Owner\n- {p['owner'] or 'Unassigned'}"
    )
    charter_deliverables = "## Deliverables\n\n" + "\n".join(
        f"- {m['name']}" for m in ms[:12]
    ) if ms else "## Deliverables\n\n(Sourced from milestones — none recorded.)"

    # Pre-format milestone INSERT values so the f-string stays manageable.
    if ms:
        milestone_values = []
        for i, m in enumerate(ms):
            mname = m["name"]
            mdesc_parts = [
                f"Phase: {m['phase'] or '—'}",
                f"Owner: {m['owner'] or '—'}",
            ]
            if m["start"]:
                mdesc_parts.insert(0, f"Start: {m['start'].strftime('%Y-%m-%d') if isinstance(m['start'], datetime) else m['start']}")
            if m["comments"]:
                mdesc_parts.append(f"Comments: {m['comments']}")
            if m["risks"]:
                mdesc_parts.append(f"Risks: {m['risks']}")
            mdesc = "\n".join(mdesc_parts)
            mstatus = map_status(m["status"])
            mrag = RAG_BY_STATUS[mstatus]
            mstart = fmt_date(m["start"])
            mdue = fmt_date(m["target"])
            milestone_values.append(
                f"  (v_pid, {sql_str(mname)}, {sql_str(mdesc)}, {mstart}, {mdue}, "
                f"{sql_str(mstatus)}, 'P2', {sql_str(mrag)}, {i})"
            )
        milestones_sql = (
            "    INSERT INTO pmo_milestones (project_id, name, description, start_date, due_date, status, priority, rag, \"order\") VALUES\n"
            + ",\n".join(milestone_values) + ";"
        )
        total_milestones += len(ms)
    else:
        milestones_sql = "    -- (no milestones in tracker for this project)"

    out_lines.append(
        f"-- ── {inserted_count + 1}. {name} ({len(ms)} milestones) ──────────────────────\n"
        f"DO $$\n"
        f"DECLARE v_pid INT; v_cid INT;\n"
        f"BEGIN\n"
        f"  IF NOT EXISTS (SELECT 1 FROM pmo_projects WHERE name = {sql_str(name)}) THEN\n"
        f"    INSERT INTO pmo_projects (\n"
        f"      name, description, status, stage, rag_status,\n"
        f"      start_date, end_date, progress\n"
        f"    ) VALUES (\n"
        f"      {sql_str(name)},\n"
        f"      {sql_str(description)},\n"
        f"      'active', 'execution', {sql_str(proj_rag)},\n"
        f"      {fmt_date(p['planned_start'])}, {fmt_date(p['planned_end'])}, {progress}\n"
        f"    ) RETURNING id INTO v_pid;\n"
        f"\n"
        f"    INSERT INTO pmo_charters (\n"
        f"      title, description, scope, deliverables, status, submitted_by_id, project_id\n"
        f"    ) VALUES (\n"
        f"      {sql_str(name)},\n"
        f"      {sql_str(description)},\n"
        f"      {sql_str(charter_scope)},\n"
        f"      {sql_str(charter_deliverables)},\n"
        f"      'active', 0, v_pid\n"
        f"    ) RETURNING id INTO v_cid;\n"
        f"\n"
        f"    UPDATE pmo_projects SET charter_id = v_cid WHERE id = v_pid;\n"
        f"\n"
        f"{milestones_sql}\n"
        f"    RAISE NOTICE 'Imported: % (pid=%, % milestones)', {sql_str(name)}, v_pid, {len(ms)};\n"
        f"  ELSE\n"
        f"    RAISE NOTICE 'Skipped (already exists): %', {sql_str(name)};\n"
        f"  END IF;\n"
        f"END $$;\n"
    )
    inserted_count += 1

out_lines.append(
    "\n-- Summary (printed by psql at end via SELECT) ──────────────────────────────\n"
    "SELECT COUNT(*) AS total_projects FROM pmo_projects;\n"
    "SELECT COUNT(*) AS total_milestones FROM pmo_milestones;\n"
    "\nCOMMIT;\n"
)

OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
OUT_PATH.write_text("\n".join(out_lines))

print(f"✓ Generated {OUT_PATH}")
print(f"  Projects: {inserted_count}")
print(f"  Milestones across all projects: {total_milestones}")
print()
print("Apply with:")
print(f"  psql \"$DATABASE_URL\" -f {OUT_PATH.relative_to(ROOT)}")
